import openpyxl
import sys

def dump_all_sheets(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_val = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"\n===== SHEET: {sheet_name} =====")
            ws = wb[sheet_name]
            ws_val = wb_val[sheet_name]

            # Dump first 40 rows, first 10 columns
            for row in range(1, 41):
                row_data = []
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell_val = ws_val.cell(row=row, column=col)
                    
                    content = ""
                    if cell.value is not None:
                        if isinstance(cell.value, str) and str(cell.value).startswith('='):
                            content = f"[{cell.value} (Val: {cell_val.value})]"
                        else:
                            content = str(cell.value)
                    row_data.append(content)
                
                if any(c != "" for c in row_data):
                    print(f"R{row}: {' | '.join(row_data)}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    dump_all_sheets("BAREME.xlsm")
