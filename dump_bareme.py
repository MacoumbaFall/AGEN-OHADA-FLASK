import openpyxl
import sys

def dump_sheet(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found. Available: {wb.sheetnames}")
            return
        
        ws = wb[sheet_name]
        wb_val = openpyxl.load_workbook(file_path, data_only=True)
        ws_val = wb_val[sheet_name]

        print(f"--- {sheet_name} ---")
        for row in range(1, 50):
            row_data = []
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell_val = ws_val.cell(row=row, column=col)
                content = f"[{cell.value} (Val: {cell_val.value})]" if cell.value is not None else ""
                row_data.append(content)
            if any(c != "" for c in row_data):
                print(f"R{row}: {' | '.join(row_data)}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    dump_sheet("BAREME.xlsm", "S.A & CA")
