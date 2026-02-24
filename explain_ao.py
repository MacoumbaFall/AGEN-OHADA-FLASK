import openpyxl

def explain_ao_divers():
    wb = openpyxl.load_workbook('BAREME.xlsm', data_only=True)
    if 'A.O' not in wb.sheetnames:
        print("Sheet A.O not found")
        return
        
    ws = wb['A.O']
    print("--- A.O Sheet Debug ---")
    for r in range(1, 50):
        row = [str(c.value) if c.value is not None else "" for c in ws[r]]
        if any(row):
            print(f"R{r}: {' | '.join(row)}")

if __name__ == "__main__":
    explain_ao_divers()
