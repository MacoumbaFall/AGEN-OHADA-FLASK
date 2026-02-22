import openpyxl

def dump_sheets(file_path, sheets):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_val = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found.")
            continue
            
        print(f"\n--- {sheet_name} ---")
        ws = wb[sheet_name]
        ws_val = wb_val[sheet_name]
        
        for row in range(1, 60):
            row_data = []
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell_val = ws_val.cell(row=row, column=col)
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        row_data.append(f"[{cell.value} (Val: {cell_val.value})]")
                    else:
                        row_data.append(str(cell.value))
                else:
                    row_data.append("")
            if any(c != "" for c in row_data):
                print(f"R{row}: {' | '.join(row_data)}")

if __name__ == "__main__":
    dump_sheets("BAREME.xlsm", ["VTE ADJUDCAT°", "AUGMT°SARL"])
